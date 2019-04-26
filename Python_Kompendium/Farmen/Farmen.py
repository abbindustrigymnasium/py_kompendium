import openpyxl
wb = openpyxl.load_workbook('Farmen.xlsx')
sheets = wb.get_sheet_names()
for i in sheets:
    i = wb.get_sheet_by_name(i)
    print(i)
    for q in range(2, i.max_row):
        induvidialinfo = [i.cell(row=q, column=w).value for w in range(1, i.max_column+1)]
        if(len(induvidialinfo) == 5): print(induvidialinfo[0], "är", induvidialinfo[1], "år gammal och bor i", induvidialinfo[2], "där hens sysselsättning är", induvidialinfo[3])
        elif(len(induvidialinfo) == 6): print("Vecka", induvidialinfo[0], "var", induvidialinfo[1], "storbonde med sinna två kämpar", induvidialinfo[2], "och", induvidialinfo[3], "deras tvekamp var", induvidialinfo[4], "och tyvärr blev", induvidialinfo[5], "hemskickad")
        elif(len(induvidialinfo) == 9): print("Vecka", induvidialinfo[0], "kördes avsnitten", induvidialinfo[1], "under", induvidialinfo[2], "där", induvidialinfo[7], "tusen personer kollade sammanlaggt")