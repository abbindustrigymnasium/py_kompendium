import openpyxl # importerar openpyxl som är ett 
wb = openpyxl.load_workbook('Farmen.xlsx') # tar in xlsx filen och döper den till workbook
sheets = wb.get_sheet_names() # tar alla sheets ur workbooken
for i in sheets: # för varje object i arrayen sheets
    i = wb.get_sheet_by_name(i) # sätter vi att i är sheetet
    print(i) # skriver ut sheetet för att veta vilken vi använder
    for q in range(2, i.max_row): # vi tar alla object i sheetet radvis förutom den första raden
        induvidialinfo = [i.cell(row=q, column=w).value for w in range(1, i.max_column+1)] # sätter induvidual info till alla object i raden vi konstanterade ovan
        if(len(induvidialinfo) == 5): print(induvidialinfo[0], "är", induvidialinfo[1], "år gammal och bor i", induvidialinfo[2], "där hens sysselsättning är", induvidialinfo[3]) # kollar arrayens sorlek och kan då skriva ut information snyggt genom att använda arrayen vi nyss tog
        elif(len(induvidialinfo) == 6): print("Vecka", induvidialinfo[0], "var", induvidialinfo[1], "storbonde med sinna två kämpar", induvidialinfo[2], "och", induvidialinfo[3], "deras tvekamp var", induvidialinfo[4], "och tyvärr blev", induvidialinfo[5], "hemskickad")
        elif(len(induvidialinfo) == 9): print("Vecka", induvidialinfo[0], "kördes avsnitten", induvidialinfo[1], "under", induvidialinfo[2], "där", induvidialinfo[7], "tusen personer kollade sammanlaggt")